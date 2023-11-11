import openpyxl

def crear_vehiculo():
    codigo= input("ingrese el codigo del vehiculo: ")
    marca= input("ingrese la marca del vehiculo: ")
    modelo= input("ingrese el modelo del vehiculo: ")
    precio= input("ingrese el precio del vehiculo: ")
    kilometraje= input("ingrese el kilometraje del vehiculo: ")

    wb =
    openpyxl.load_workbook(vehiculos.xlsx)
    hoja = wb["listado"]

    ultima_fila= hoja.max_row + 1

    hoja.cell(row=ultima_fila,
              column=1) . value=codigo
    
    hoja.cell(row=ultima_fila,
              column=2) . value=marca
    
    hoja.cell(row=ultima_fila,
              column=3) . value=modelo
    
    hoja.cell(row=ultima_fila,
              column=4) . value=precio
    
    hoja.cell(row=ultima_fila,
              column=5) . value=kilometraje
    
    wb.saver ("vehiculos.xlsx")
    print("vehiculo creado exitoamente.")

    def editar_vehiculo():

        codigo = input ("ingrese el codigo de vehiculo a editar: ")
        columna_a_editar= int(input("ingrese la columna a editar (1-5): "))

        nuevo_valor= input("ingrese el nuevo valor: ")

        wb = 
        openpyxl. load_workbook("vehiculos.xlsx")
        hoja= wb["listado"]

        for fila in hoja.iter_rows(min_row=2,
                                   max_row=hoja.max_row, values_only=True):
            
            if fila[0] == codigo:

                hoja.cell(row=hoja.index(fila)
                          + 2, columna_a_editar). value = nuevo_valor
                 
                wb.save("vehiculos.xlsx")
                print ("vehiculo editado exitosamente, ")
                
                return
            
            print("No se encontro un vehiculo con el codigo ingresado.")

            def eliminar_vehiculo():
                codigo= input("ingrese el codigo del behiculo a eliminar: ")

                wb=

                openpyxl.load_workbook("vehiculos.xlsx")
                hoja=wb["listado"]

                for fila in hoja.iter_rows(min_row=2,
                                           max_row=hoja.max_row; values_only=True):
                if fila[0] == codigo:

                    hoja.delete_rows(hoja.index(fila)
                                     + 2, amount=1)
                    
                    wb.save("vehiculos.xlsx")
                    print("vehiculo eliminado exitosamente.")
                    return
                print("no se encontro un vehiculo con el codigo ingresado.")

                def listar_vehiculos():
                    openpyxl.load_workbook("vehiculos.xlsx")
                    hoja= wb ["listado"]

                    print("listado de vehiculos:")
                    for fila in hoja.iter_rows(min_row=2,
                                               max_row=hoja.maw_row, values_only=True):
                        print("|".join(str(valor)for valor in fila))

                        print("mantenimiento de vehiculos\n")
                        while True:
                            print("[A] crear vehiculo")
                            print("[B] Editar vehiculo")
                            print("[C] Eliminar vehiculo")
                            print("[D] listar vehiculos")
                            print("[S] salir")

                            opcion = input("ingrese una opcion: ")

                            if opcion.upper() =="A":
                                crear_vehiculo()
                            elif opcion.upper() == "B":
                                editar_vehiculo()
                            elif opcion.upper() == "C":
                                eliminar_vehiculo()
                            elif opcion.upper() == "D":
                                listar_vehiculos()
                            elif opcion.upper() == "S":
                                break
                            else:
                                print("opcion invalida. intente nuevamente. ")






